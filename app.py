import os
from io import BytesIO
from datetime import datetime, timedelta
from random import uniform, randint

from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, flash
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import desc, inspect, text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "iv-monitoring-system-secret")

DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///iv_monitor.db")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

MONITOR_OPTIONS = ["Fariz", "Hareny", "Madam Ku Lee Chin"]
DEFAULT_API_KEY = os.getenv("API_KEY", "IVMONITOR123")
DROP_FACTOR = float(os.getenv("DROP_FACTOR", "20"))  # 20 drops/ml macrodrip default
AUTO_DEMO = os.getenv("AUTO_DEMO", "true").lower() not in ["0", "false", "no"]


class MonitorLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    monitor_name = db.Column(db.String(100), nullable=False)
    login_time = db.Column(db.DateTime, default=datetime.now, nullable=False)


class PatientSlot(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    patient_name = db.Column(db.String(120), nullable=False)
    patient_code = db.Column(db.String(30), nullable=False, default="PT001A")
    ward_number = db.Column(db.String(50), nullable=False, default="Ward 3A")
    bed_number = db.Column(db.String(50), nullable=False, default="Bed 05")
    full_weight_g = db.Column(db.Float, nullable=False, default=800.0)
    empty_weight_g = db.Column(db.Float, nullable=False, default=50.0)
    current_weight_g = db.Column(db.Float, nullable=False, default=540.0)
    remaining_ml = db.Column(db.Float, nullable=False, default=340.0)
    current_level_percent = db.Column(db.Float, nullable=False, default=68.0)
    current_drop_rate = db.Column(db.Float, nullable=False, default=24.0)
    current_flow_rate_ml_hr = db.Column(db.Float, nullable=False, default=72.0)
    current_status = db.Column(db.String(20), nullable=False, default="Normal")
    last_update_time = db.Column(db.DateTime, default=datetime.now, nullable=False)
    last_demo_time = db.Column(db.DateTime, nullable=True)


class Reading(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    patient_id = db.Column(db.Integer, db.ForeignKey("patient_slot.id"), nullable=False)
    weight_g = db.Column(db.Float, nullable=False)
    remaining_ml = db.Column(db.Float, nullable=False, default=0.0)
    level_percent = db.Column(db.Float, nullable=False)
    drops_per_min = db.Column(db.Float, nullable=False, default=0.0)
    flow_rate_ml_hr = db.Column(db.Float, nullable=False, default=0.0)
    status = db.Column(db.String(20), nullable=False, default="Normal")
    source = db.Column(db.String(30), nullable=False, default="system")
    created_at = db.Column(db.DateTime, default=datetime.now, nullable=False)


class Alert(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    patient_id = db.Column(db.Integer, db.ForeignKey("patient_slot.id"), nullable=False)
    alert_type = db.Column(db.String(50), nullable=False)
    priority = db.Column(db.String(20), nullable=False, default="Info")
    message = db.Column(db.String(255), nullable=False)
    level_percent = db.Column(db.Float, nullable=False, default=0.0)
    status = db.Column(db.String(30), nullable=False, default="New")
    acknowledged_by = db.Column(db.String(100), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now, nullable=False)


def now_local():
    return datetime.now()


def fmt_time(dt=None):
    dt = dt or now_local()
    return dt.strftime("%I:%M:%S %p")


def fmt_time_short(dt=None):
    dt = dt or now_local()
    return dt.strftime("%I:%M %p")


def fmt_date(dt=None):
    dt = dt or now_local()
    return dt.strftime("%d %B %Y (%a)")


def calculate_flow(drops_per_min):
    try:
        return round((float(drops_per_min) * 60.0) / max(DROP_FACTOR, 1.0), 2)
    except Exception:
        return 0.0


def calculate_level(weight_g, full_weight_g):
    try:
        percent = (float(weight_g) / max(float(full_weight_g), 1.0)) * 100
    except Exception:
        percent = 0
    return round(max(0, min(100, percent)), 2)


def status_from_level(level):
    if level <= 10:
        return "Critical"
    if level <= 30:
        return "Low"
    return "Normal"


def remaining_ml_from_weight(weight_g, full_weight_g):
    # Approximation for a 500 ml IV bag. The dashboard still displays total weight separately.
    try:
        return round(max(0, min(500, (float(weight_g) / max(float(full_weight_g), 1.0)) * 500)), 0)
    except Exception:
        return 0


def safe_float(value, default=0.0):
    try:
        return float(value)
    except Exception:
        return default


def ensure_schema_upgrades():
    inspector = inspect(db.engine)
    tables = set(inspector.get_table_names())
    if "patient_slot" in tables:
        existing = {c["name"] for c in inspector.get_columns("patient_slot")}
        upgrades = {
            "patient_code": "VARCHAR(30) DEFAULT 'PT001A'",
            "remaining_ml": "FLOAT DEFAULT 0",
            "last_demo_time": "TIMESTAMP",
        }
        for col, ddl in upgrades.items():
            if col not in existing:
                db.session.execute(text(f"ALTER TABLE patient_slot ADD COLUMN {col} {ddl}"))
    if "reading" in tables:
        existing = {c["name"] for c in inspector.get_columns("reading")}
        if "remaining_ml" not in existing:
            db.session.execute(text("ALTER TABLE reading ADD COLUMN remaining_ml FLOAT DEFAULT 0"))
    if "alert" in tables:
        existing = {c["name"] for c in inspector.get_columns("alert")}
        upgrades = {
            "priority": "VARCHAR(20) DEFAULT 'Info'",
            "status": "VARCHAR(30) DEFAULT 'New'",
            "acknowledged_by": "VARCHAR(100)",
        }
        for col, ddl in upgrades.items():
            if col not in existing:
                db.session.execute(text(f"ALTER TABLE alert ADD COLUMN {col} {ddl}"))
    db.session.commit()


def seed_database():
    if PatientSlot.query.count() == 0:
        p1 = PatientSlot(
            patient_name="Patient A",
            patient_code="PT001A",
            ward_number="Ward 3A",
            bed_number="Bed 05",
            full_weight_g=794.0,
            empty_weight_g=50.0,
            current_weight_g=540.0,
            remaining_ml=340.0,
            current_level_percent=68.0,
            current_drop_rate=24.0,
            current_flow_rate_ml_hr=72.0,
            current_status="Normal",
        )
        p2 = PatientSlot(
            patient_name="Patient B",
            patient_code="PT002B",
            ward_number="Ward 3A",
            bed_number="Bed 06",
            full_weight_g=955.0,
            empty_weight_g=50.0,
            current_weight_g=210.0,
            remaining_ml=110.0,
            current_level_percent=22.0,
            current_drop_rate=12.0,
            current_flow_rate_ml_hr=36.0,
            current_status="Low",
        )
        db.session.add_all([p1, p2])
        db.session.commit()

    patients = PatientSlot.query.order_by(PatientSlot.id).limit(2).all()
    if len(patients) >= 2:
        defaults = [
            ("Patient A", "PT001A", "Bed 05", 794.0, 540.0, 24),
            ("Patient B", "PT002B", "Bed 06", 955.0, 210.0, 12),
        ]
        for patient, default in zip(patients, defaults):
            if not patient.patient_code:
                patient.patient_code = default[1]
            if not patient.ward_number:
                patient.ward_number = "Ward 3A"
            if not patient.bed_number:
                patient.bed_number = default[2]
            if not patient.full_weight_g:
                patient.full_weight_g = default[3]
            if patient.current_drop_rate is None:
                patient.current_drop_rate = default[5]
        db.session.commit()

    for idx, patient in enumerate(PatientSlot.query.order_by(PatientSlot.id).limit(2).all()):
        if Reading.query.filter_by(patient_id=patient.id).count() == 0:
            base = now_local() - timedelta(minutes=64)
            if idx == 0:
                weights = [595, 572, 558, 546, 530, 512, 500, 486, 470, 455, 438, 424, 402, 388, 370, 358, 340, 326, 310, 292, 276, 260, 246, 232, 218]
                drops =  [28,  26,  30,  25,  27,  26,  24,  25,  24,  28,  25,  24,  26,  23,  25,  22,  24,  23,  22,  20,  21,  20,  19,  20,  18]
            else:
                weights = [520, 498, 475, 452, 430, 405, 370, 350, 330, 302, 282, 262, 240, 218, 210, 196, 178, 164, 150, 138, 126, 118, 112, 106, 100]
                drops =  [15,  14,  13,  12,  13,  11,  12,  11,  10,  12,  10,  9,   11,  9,   8,   9,   7,   8,   7,   7,   8,   6,   7,   7,   6]
            for n, (weight, drop) in enumerate(zip(weights, drops)):
                created = base + timedelta(minutes=n * 2.5)
                level = calculate_level(weight, patient.full_weight_g)
                flow = calculate_flow(drop)
                status = status_from_level(level)
                remaining = remaining_ml_from_weight(weight, patient.full_weight_g)
                db.session.add(Reading(
                    patient_id=patient.id,
                    weight_g=weight,
                    remaining_ml=remaining,
                    level_percent=level,
                    drops_per_min=drop,
                    flow_rate_ml_hr=flow,
                    status=status,
                    source="demo",
                    created_at=created,
                ))
            latest = Reading.query.filter_by(patient_id=patient.id).order_by(desc(Reading.created_at)).first()
            if latest:
                patient.current_weight_g = latest.weight_g
                patient.remaining_ml = latest.remaining_ml
                patient.current_level_percent = latest.level_percent
                patient.current_drop_rate = latest.drops_per_min
                patient.current_flow_rate_ml_hr = latest.flow_rate_ml_hr
                patient.current_status = latest.status
                patient.last_update_time = latest.created_at
        db.session.commit()

    if Alert.query.count() == 0:
        p1, p2 = PatientSlot.query.order_by(PatientSlot.id).limit(2).all()
        entries = [
            (p2.id, "Low IV Level", "Critical", "Patient B IV level below 25%. Please monitor.", 22, "New", 0),
            (p2.id, "Nearing Empty", "Warning", "Patient B IV level is below 20%. Refill soon.", 18, "New", 4),
            (p1.id, "Stable", "Info", "Patient A IV level is normal. No action required.", 68, "Acknowledged", 6),
            (p2.id, "Flow Rate High", "Warning", "Flow rate above normal threshold.", 30, "New", 34),
            (p1.id, "Low IV Level", "Warning", "IV level below 30%. Please monitor.", 29, "Acknowledged", 54),
            (p1.id, "Air In Line", "Warning", "Air detected in line. Check connection.", 61, "Resolved", 118),
        ]
        for patient_id, alert_type, priority, message, level, status, minutes_ago in entries:
            db.session.add(Alert(
                patient_id=patient_id,
                alert_type=alert_type,
                priority=priority,
                message=message,
                level_percent=level,
                status=status,
                created_at=now_local() - timedelta(minutes=minutes_ago),
            ))
        db.session.commit()


def create_alert(patient):
    if patient.current_status == "Normal":
        return
    latest = Alert.query.filter_by(patient_id=patient.id, alert_type="Low IV Level").order_by(desc(Alert.created_at)).first()
    if latest and (now_local() - latest.created_at).total_seconds() < 300:
        return
    priority = "Critical" if patient.current_status == "Critical" else "Warning"
    msg = f"{patient.patient_name} IV level is {patient.current_status.lower()} at {patient.current_level_percent:.0f}%. Please monitor."
    db.session.add(Alert(
        patient_id=patient.id,
        alert_type="Low IV Level",
        priority=priority,
        message=msg,
        level_percent=patient.current_level_percent,
        status="New",
        created_at=now_local(),
    ))


def save_reading(patient, weight_g, drops_per_min, source="esp32"):
    weight_g = abs(safe_float(weight_g, patient.current_weight_g))
    if source == "esp32" and weight_g <= 0:
        weight_g = max(patient.current_weight_g or 0, 1)
    drops_per_min = max(0, safe_float(drops_per_min, patient.current_drop_rate or 0))
    level = calculate_level(weight_g, patient.full_weight_g)
    remaining = remaining_ml_from_weight(weight_g, patient.full_weight_g)
    flow = calculate_flow(drops_per_min)
    status = status_from_level(level)
    timestamp = now_local()

    patient.current_weight_g = round(weight_g, 2)
    patient.remaining_ml = remaining
    patient.current_level_percent = level
    patient.current_drop_rate = drops_per_min
    patient.current_flow_rate_ml_hr = flow
    patient.current_status = status
    patient.last_update_time = timestamp

    db.session.add(Reading(
        patient_id=patient.id,
        weight_g=round(weight_g, 2),
        remaining_ml=remaining,
        level_percent=level,
        drops_per_min=drops_per_min,
        flow_rate_ml_hr=flow,
        status=status,
        source=source,
        created_at=timestamp,
    ))
    create_alert(patient)
    db.session.commit()


def simulate_live_if_needed():
    if not AUTO_DEMO:
        return
    current = now_local()
    patients = PatientSlot.query.order_by(PatientSlot.id).limit(2).all()
    for idx, patient in enumerate(patients):
        # If ESP32 is sending data, let real readings take priority.
        latest_esp32 = Reading.query.filter_by(patient_id=patient.id, source="esp32").order_by(desc(Reading.created_at)).first()
        if latest_esp32 and (current - latest_esp32.created_at).total_seconds() < 20:
            continue
        if patient.last_demo_time and (current - patient.last_demo_time).total_seconds() < 5:
            continue
        decrement = uniform(1.5, 4.0) if idx == 0 else uniform(0.8, 2.5)
        new_weight = max(10, (patient.current_weight_g or 0) - decrement)
        base_drop = 24 if idx == 0 else 12
        drops = max(0, base_drop + randint(-2, 2))
        patient.last_demo_time = current
        save_reading(patient, new_weight, drops, source="demo-live")


def get_readings(patient_id, limit=30):
    return Reading.query.filter_by(patient_id=patient_id).order_by(desc(Reading.created_at)).limit(limit).all()[::-1]


def patient_payload(patient):
    readings = get_readings(patient.id, 30)
    return {
        "id": patient.id,
        "patient_name": patient.patient_name,
        "patient_code": patient.patient_code or f"PT{patient.id:03d}",
        "ward_number": patient.ward_number,
        "bed_number": patient.bed_number,
        "full_weight_g": round(patient.full_weight_g or 0, 1),
        "empty_weight_g": round(patient.empty_weight_g or 0, 1),
        "current_weight_g": round(patient.current_weight_g or 0, 1),
        "remaining_ml": round(patient.remaining_ml or 0, 0),
        "current_level_percent": round(patient.current_level_percent or 0, 1),
        "current_drop_rate": round(patient.current_drop_rate or 0, 1),
        "current_flow_rate_ml_hr": round(patient.current_flow_rate_ml_hr or 0, 1),
        "current_status": patient.current_status or "Normal",
        "last_update_time": fmt_time(patient.last_update_time),
        "last_update_date": fmt_date(patient.last_update_time),
        "readings": [
            {
                "label": r.created_at.strftime("%H:%M"),
                "time": fmt_time(r.created_at),
                "weight_g": round(r.weight_g or 0, 1),
                "remaining_ml": round(r.remaining_ml or 0, 0),
                "level_percent": round(r.level_percent or 0, 1),
                "drops_per_min": round(r.drops_per_min or 0, 1),
                "flow_rate_ml_hr": round(r.flow_rate_ml_hr or 0, 1),
                "status": r.status,
                "source": r.source,
            }
            for r in readings
        ],
    }


def alert_payload(alert):
    patient = PatientSlot.query.get(alert.patient_id)
    return {
        "id": alert.id,
        "time": fmt_time_short(alert.created_at),
        "date": alert.created_at.strftime("%d %b %Y"),
        "patient_id": alert.patient_id,
        "patient_name": patient.patient_name if patient else f"Patient {alert.patient_id}",
        "patient_code": patient.patient_code if patient else "-",
        "ward": patient.ward_number if patient else "-",
        "bed": patient.bed_number if patient else "-",
        "alert_type": alert.alert_type,
        "priority": alert.priority,
        "message": alert.message,
        "level_percent": round(alert.level_percent or 0, 1),
        "status": alert.status,
    }


def build_dashboard_payload():
    simulate_live_if_needed()
    patients = PatientSlot.query.order_by(PatientSlot.id).limit(2).all()
    patient_items = [patient_payload(p) for p in patients]
    alerts = Alert.query.order_by(desc(Alert.created_at)).limit(12).all()
    longest = max(patient_items, key=lambda p: len(p["readings"]), default={"readings": []})
    labels = [r["label"] for r in longest["readings"]]
    active_alerts = [a for a in alerts if a.status == "New"]
    return {
        "server_time": fmt_time(),
        "server_date": fmt_date(),
        "patients": patient_items,
        "alerts": [alert_payload(a) for a in alerts],
        "active_alert_count": len(active_alerts),
        "drop_comparison": {
            "labels": labels,
            "series": [
                {
                    "name": p["patient_name"],
                    "patient_id": p["id"],
                    "data": [r["drops_per_min"] for r in p["readings"]],
                }
                for p in patient_items
            ],
        },
        "system": {
            "data_source": "PostgreSQL (Render Cloud)" if DATABASE_URL.startswith("postgresql") else "SQLite Local Test",
            "connected": True,
            "uptime": "2d 14h 35m 12s",
            "last_backup": "22 May 2024, 02:00 AM",
            "auto_demo": AUTO_DEMO,
        }
    }


@app.before_request
def ready_database():
    db.create_all()
    ensure_schema_upgrades()
    seed_database()


@app.context_processor
def inject_globals():
    return {
        "monitor_options": MONITOR_OPTIONS,
        "active_language": session.get("language", "en"),
        "current_monitor": session.get("monitor_name", "Fariz"),
        "dashboard_data": build_dashboard_payload,
    }


@app.route("/")
def index():
    return render_template("landing.html")


@app.route("/select-monitor")
def select_monitor():
    return render_template("select_monitor.html")


@app.route("/login", methods=["POST"])
def login():
    monitor_name = request.form.get("monitor_name", "").strip()
    if monitor_name not in MONITOR_OPTIONS:
        flash("Please select a valid monitor.")
        return redirect(url_for("select_monitor"))
    session["monitor_name"] = monitor_name
    db.session.add(MonitorLog(monitor_name=monitor_name, login_time=now_local()))
    db.session.commit()
    return redirect(url_for("dashboard"))


@app.route("/dashboard")
def dashboard():
    if "monitor_name" not in session:
        session["monitor_name"] = "Fariz"
    data = build_dashboard_payload()
    return render_template("dashboard.html", data=data)


@app.route("/set-language/<lang>", methods=["POST"])
def set_language(lang):
    if lang in ["en", "ms"]:
        session["language"] = lang
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/api/dashboard-data")
def api_dashboard_data():
    return jsonify(build_dashboard_payload())


@app.route("/api/update", methods=["POST"])
def api_update():
    payload = request.get_json(silent=True) or request.form
    api_key = payload.get("api_key")
    if api_key != DEFAULT_API_KEY:
        return jsonify({"success": False, "message": "Invalid API key"}), 401
    patient_id = int(payload.get("patient_id", 1))
    patient = PatientSlot.query.get(patient_id)
    if not patient:
        return jsonify({"success": False, "message": "Patient not found"}), 404
    weight = payload.get("weight_g") or payload.get("weight") or payload.get("total_weight") or payload.get("total_weight_g")
    if weight is None:
        return jsonify({"success": False, "message": "weight_g / weight / total_weight_g is required"}), 400
    drops = payload.get("drops_per_min") or payload.get("drop_rate") or payload.get("drip_rate") or patient.current_drop_rate or 0
    save_reading(patient, weight, drops, source="esp32")
    return jsonify({"success": True, "patient": patient_payload(patient)})


@app.route("/manual-reading/<int:patient_id>", methods=["POST"])
def manual_reading(patient_id):
    patient = PatientSlot.query.get_or_404(patient_id)
    save_reading(patient, request.form.get("weight_g"), request.form.get("drops_per_min"), source="manual")
    return redirect(url_for("dashboard") + "#monitors")


@app.route("/update-patient/<int:patient_id>", methods=["POST"])
def update_patient(patient_id):
    patient = PatientSlot.query.get_or_404(patient_id)
    patient.patient_name = request.form.get("patient_name", patient.patient_name).strip() or patient.patient_name
    patient.patient_code = request.form.get("patient_code", patient.patient_code).strip() or patient.patient_code
    patient.ward_number = request.form.get("ward_number", patient.ward_number).strip() or patient.ward_number
    patient.bed_number = request.form.get("bed_number", patient.bed_number).strip() or patient.bed_number
    patient.full_weight_g = safe_float(request.form.get("full_weight_g"), patient.full_weight_g)
    patient.empty_weight_g = safe_float(request.form.get("empty_weight_g"), patient.empty_weight_g)
    patient.current_level_percent = calculate_level(patient.current_weight_g, patient.full_weight_g)
    patient.remaining_ml = remaining_ml_from_weight(patient.current_weight_g, patient.full_weight_g)
    patient.current_status = status_from_level(patient.current_level_percent)
    db.session.commit()
    return redirect(url_for("dashboard") + "#patients")


@app.route("/acknowledge-alert/<int:alert_id>", methods=["POST"])
def acknowledge_alert(alert_id):
    alert = Alert.query.get_or_404(alert_id)
    alert.status = "Acknowledged"
    alert.acknowledged_by = session.get("monitor_name", "Staff")
    db.session.commit()
    return redirect(url_for("dashboard") + "#alerts")


@app.route("/export/excel")
def export_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Current Status"
    headers = ["Patient Code", "Patient Name", "Ward", "Bed", "Total Weight (g)", "Remaining (ml)", "IV Level (%)", "Drip Rate", "Flow Rate", "Status", "Last Update"]
    ws.append(headers)
    for p in PatientSlot.query.order_by(PatientSlot.id).limit(2).all():
        ws.append([p.patient_code, p.patient_name, p.ward_number, p.bed_number, p.current_weight_g, p.remaining_ml, p.current_level_percent, p.current_drop_rate, p.current_flow_rate_ml_hr, p.current_status, fmt_time(p.last_update_time)])

    ws2 = wb.create_sheet("Readings")
    ws2.append(["Date", "Time", "Patient", "Weight (g)", "Remaining (ml)", "IV Level (%)", "Drops/min", "Flow Rate (ml/hr)", "Status", "Source"])
    for r in Reading.query.order_by(desc(Reading.created_at)).limit(500).all():
        p = PatientSlot.query.get(r.patient_id)
        ws2.append([r.created_at.strftime("%d/%m/%Y"), fmt_time(r.created_at), p.patient_name if p else r.patient_id, r.weight_g, r.remaining_ml, r.level_percent, r.drops_per_min, r.flow_rate_ml_hr, r.status, r.source])

    ws3 = wb.create_sheet("Alerts")
    ws3.append(["Date", "Time", "Patient", "Alert Type", "Priority", "Message", "Status", "Level (%)"])
    for a in Alert.query.order_by(desc(Alert.created_at)).all():
        p = PatientSlot.query.get(a.patient_id)
        ws3.append([a.created_at.strftime("%d/%m/%Y"), fmt_time(a.created_at), p.patient_name if p else a.patient_id, a.alert_type, a.priority, a.message, a.status, a.level_percent])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="083B77")
            cell.alignment = Alignment(horizontal="center")
        for col in sheet.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 38)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"iv_monitoring_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", 5000)), debug=True)
